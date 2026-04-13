from __future__ import annotations

import re
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from enhance_workbook_low_risk import (
    BLUE,
    GRAY,
    INT_FMT,
    MONEY_FMT,
    NAVY,
    PURPLE,
    SCENARIO_RESULTS,
    TEAL,
    WHITE,
    WORKBOOK,
    replace_defined_name,
    safe_merge,
    set_header,
    set_label,
    thin_bottom,
)


ANALYSIS_SHEET = "Scenario Analysis"
PIVOT_SHEET = "Scenario Pivot Lab"
IC_SHEET = "IC Switch Scenarios"
PM_SHEET = "PM After Switch"
IC_HELPER_FIRST_ROW = 36
IC_HELPER_LAST_ROW = 170
PM_HELPER_FIRST_ROW = 51
PM_HELPER_LAST_ROW = 350
RESULT_FIRST_ROW = 5
TEXT = "1C2B3C"
LINK = "2F5496"
NOTE = "F7F9FC"
BORDER = "A6A6A6"


def iter_real_cells(ws, cell_range: str):
    for row in ws[cell_range]:
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            yield cell


def style_cells(
    ws,
    cell_range: str,
    *,
    fill: str | None = None,
    font_color: str = TEXT,
    bold: bool = False,
    size: int = 10,
    h_align: str = "left",
    v_align: str = "center",
    wrap: bool = True,
) -> None:
    for cell in iter_real_cells(ws, cell_range):
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", color=font_color, bold=bold, size=size)
        cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)


def set_link(ws, cell_ref: str, label: str, target_sheet: str) -> None:
    ws[cell_ref] = label
    ws[cell_ref].hyperlink = f"#'{target_sheet}'!A1"
    ws[cell_ref].style = "Hyperlink"
    ws[cell_ref].font = Font(name="Aptos", color=LINK, underline="single")
    ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")


def get_result_last_row(wb) -> int:
    table = wb[SCENARIO_RESULTS].tables["tblScenarioResults"]
    return int(table.ref.split(":")[1][1:])


def move_sheet_after(wb, sheet_name: str, after_sheet: str) -> None:
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    insert_at = wb.sheetnames.index(after_sheet) + 1
    wb._sheets.insert(insert_at, ws)


def clear_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font(color="000000", bold=False)
        cell.alignment = Alignment(horizontal="general", vertical="bottom")
        cell.border = cell.border.copy()


def set_section_band(ws, cell_range: str, title: str, fill: str) -> None:
    safe_merge(ws, cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE if fill == NAVY else "000000", bold=True, size=12 if fill == NAVY else 11)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_dashboard(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4472C4"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[114].height = 22

    style_cells(ws, "A1:P1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:P2", fill=TEAL, font_color=TEXT, size=10)
    for cell_range in ("A4:D4", "E4:H4", "I4:L4", "A8:C8", "D8:F8", "G8:I8", "J8:L8"):
        style_cells(ws, cell_range, fill=GRAY, font_color=TEXT, bold=True, h_align="center")
    style_cells(ws, "A13:P13", fill=NAVY, font_color=WHITE, bold=True, h_align="center")

    set_section_band(ws, "A114:P114", "Navigation", NAVY)
    ws.row_dimensions[115].height = 20
    ws.row_dimensions[116].height = 20

    nav_links = [
        ("A115", "Dashboard", "Financial Dashboard"),
        ("C115", "Model Inputs", "Model Inputs"),
        ("E115", "Savings Projection", "Savings Projection"),
        ("G115", "Scenario Lab", "Scenario Lab"),
        ("I115", "IC Switch", IC_SHEET),
        ("K115", "PM Switch", PM_SHEET),
        ("M115", "Analysis", ANALYSIS_SHEET),
        ("O115", "Scenario Results", SCENARIO_RESULTS),
        ("A116", "Tax", "Tax Assumptions"),
        ("C116", "Pivot Lab", PIVOT_SHEET),
    ]
    for cell_ref, label, target_sheet in nav_links:
        set_link(ws, cell_ref, label, target_sheet)
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def style_model_inputs_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "5B9BD5"
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36
    for row in range(4, 27):
        ws.row_dimensions[row].height = 20

    style_cells(ws, "A1:P1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:P2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:B4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "E4:H4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A21:B21", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "B5:B25", fill=BLUE, font_color=LINK, wrap=False)
    style_cells(ws, "F5:H26", fill=BLUE, font_color=LINK, wrap=False)
    style_cells(ws, "J4:P22", fill=NOTE, font_color=TEXT)
    style_cells(ws, "J24:P31", fill=NOTE, font_color=TEXT)
    set_link(ws, "A3", "Back to Financial Dashboard", "Financial Dashboard")


def style_savings_projection_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "70AD47"
    ws.freeze_panes = "A19"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36
    for row in range(4, 18):
        ws.row_dimensions[row].height = 20
    for row in range(18, 59):
        ws.row_dimensions[row].height = 18

    style_cells(ws, "A1:BM1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:BM2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:E4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "F4:N4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A18:BM18", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "B5:B17", fill=BLUE, font_color=LINK, wrap=False)
    style_cells(ws, "F5:K11", fill=NOTE)
    style_cells(ws, "M7:AM8", fill=NOTE)
    set_link(ws, "A3", "Back to Financial Dashboard", "Financial Dashboard")


def style_scenario_lab_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "A5A5A5"
    ws.freeze_panes = "A125"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    for row in range(4, 27):
        ws.row_dimensions[row].height = 20
    for row in range(31, 168):
        ws.row_dimensions[row].height = 18

    style_cells(ws, "A1:AG1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:AG2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:D4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "F4:I4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A31:AG31", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A79:AG79", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A127:AG127", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A5:A18", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "F5:F18", fill=GRAY, font_color=TEXT, bold=True)
    set_link(ws, "A3", "Back to Financial Dashboard", "Financial Dashboard")


def style_ic_switch_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "ED7D31"
    ws.freeze_panes = "A20"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 40
    for row in range(4, 19):
        ws.row_dimensions[row].height = 20
    for row in range(20, 31):
        ws.row_dimensions[row].height = 20
    for row in range(35, 171):
        ws.row_dimensions[row].height = 18

    style_cells(ws, "A1:AB1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:AB2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:Q4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A12:G12", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A20:V20", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A21:Z21", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A35:AB35", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A5:A9", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "B5:Q9", fill=NOTE, font_color=TEXT)
    style_cells(ws, "I14:L18", fill=GRAY, font_color=TEXT, bold=True)
    set_link(ws, "A3", "Back to Financial Dashboard", "Financial Dashboard")


def style_pm_switch_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FFC000"
    ws.freeze_panes = "A21"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 42
    for row in range(4, 19):
        ws.row_dimensions[row].height = 20
    for row in range(20, 42):
        ws.row_dimensions[row].height = 20
    for row in range(50, 351):
        ws.row_dimensions[row].height = 18

    style_cells(ws, "A1:AJ1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:AJ2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:W4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A12:G12", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A20:AB20", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A21:AF21", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A50:AJ50", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A5:A10", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "B5:W10", fill=NOTE, font_color=TEXT)
    style_cells(ws, "I14:L18", fill=GRAY, font_color=TEXT, bold=True)
    set_link(ws, "A3", "Back to Financial Dashboard", "Financial Dashboard")


def style_tax_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "5B9BD5"
    ws.freeze_panes = "A7"
    ws.row_dimensions[1].height = 26
    for row in range(2, 6):
        ws.row_dimensions[row].height = 22
    for row in range(7, 27):
        ws.row_dimensions[row].height = 20

    style_cells(ws, "A1:S1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:S5", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A7:S7", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A8:S8", fill=NAVY, font_color=WHITE, bold=True, h_align="center")


def configure_analysis_formula_block(ws, result_last_row: int) -> None:
    del result_last_row
    ws["K8"] = '=IFERROR(MATCH($K$7,tblScenarioResults[Result Key],0),"")'
    ws["K10"] = '=IFERROR(MATCH($K$9,tblScenarioResults[Result Key],0),"")'

    column_map = {
        14: "Path",
        15: "Scenario",
        16: "Horizon",
        17: "Cash/Gross Comp",
        18: "Taxable Liquid",
        19: "Retirement",
        20: "Liquid Net Worth",
        21: "vs $5M Cash",
        22: "vs $5M 50%",
        23: "vs $7M Cash",
        24: "vs $7M 50%",
        25: "First $5M Cash",
        26: "First $5M 50%",
        27: "First $7M Cash",
        28: "First $7M 50%",
        29: "Read",
    }
    for row, column_name in column_map.items():
        ws[f"B{row}"] = f'=IFERROR(INDEX(tblScenarioResults[{column_name}],$K$8),"")'
        ws[f"C{row}"] = f'=IFERROR(INDEX(tblScenarioResults[{column_name}],$K$10),"")'

    for row in range(34, 37):
        for col in ("B", "C", "D"):
            ws[f"{col}{row}"] = (
                '=SUMIFS('
                'tblScenarioResults[Liquid Net Worth],'
                'tblScenarioResults[Path],"IC Switch",'
                f'tblScenarioResults[Scenario],$A{row}&" "&{col}$33,'
                'tblScenarioResults[Horizon],$G$5)'
            )

    for row in range(34, 39):
        for col in ("K", "L", "M", "N"):
            ws[f"{col}{row}"] = (
                '=SUMIFS('
                'tblScenarioResults[Liquid Net Worth],'
                'tblScenarioResults[Path],"PM After Switch",'
                f'tblScenarioResults[Scenario],$J{row}&" "&{col}$33,'
                'tblScenarioResults[Horizon],$G$5)'
            )


def style_analysis_sheet(ws, result_last_row: int) -> None:
    configure_analysis_formula_block(ws, result_last_row)

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "5B9BD5"
    ws.freeze_panes = "A4"

    widths = {
        "A": 22,
        "B": 28,
        "C": 16,
        "D": 16,
        "E": 3,
        "F": 18,
        "G": 20,
        "H": 3,
        "I": 18,
        "J": 20,
        "K": 30,
        "L": 3,
        "M": 18,
        "N": 14,
        "O": 3,
        "P": 3,
        "V": 3,
        "W": 3,
        "X": 26,
        "Y": 14,
        "Z": 18,
        "AA": 18,
        "AB": 18,
        "AC": 18,
        "AD": 14,
        "AE": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col in ("J", "K", "L", "M", "N", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"):
        ws.column_dimensions[col].hidden = True

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20
    for row in range(4, 11):
        ws.row_dimensions[row].height = 22
    for row in range(12, 30):
        ws.row_dimensions[row].height = 20
    for row in range(32, 39):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[40].height = 20

    style_cells(ws, "A1:P1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:P2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:G4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A12:D12", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A13:D13", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A32:D32", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "J32:N32", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "B33:D33", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "K33:N33", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A40:G40", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "A14:A29", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "A34:A36", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "J34:J38", fill=GRAY, font_color=TEXT, bold=True)
    style_cells(ws, "J4:N10", fill=NOTE, font_color=TEXT)

    set_link(ws, "A3", "Dashboard", "Financial Dashboard")
    set_link(ws, "C3", "Scenario Results", SCENARIO_RESULTS)
    set_link(ws, "E3", "Pivot Lab", PIVOT_SHEET)

    for cell_ref in ("B5", "B6", "G5", "G6"):
        ws[cell_ref].fill = PatternFill("solid", fgColor=BLUE)
        ws[cell_ref].font = Font(name="Aptos", color=LINK)
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")
        thin_bottom(ws[cell_ref])

    for cell_ref in ("A12", "A32", "J32", "A40"):
        ws[cell_ref].font = Font(name="Aptos", color=WHITE, bold=True)


def liquid_series_formula(row: int, year_offset: int) -> str:
    return (
        f'=IF($B{row}="IC Switch",'
        f'IFERROR(INDEX(\'{IC_SHEET}\'!$X${IC_HELPER_FIRST_ROW}:$X${IC_HELPER_LAST_ROW},'
        f'MATCH($C{row},\'{IC_SHEET}\'!$A${IC_HELPER_FIRST_ROW}:$A${IC_HELPER_LAST_ROW},0)+{year_offset}),""),'
        f'IFERROR(INDEX(\'{PM_SHEET}\'!$AF${PM_HELPER_FIRST_ROW}:$AF${PM_HELPER_LAST_ROW},'
        f'MATCH($C{row},\'{PM_SHEET}\'!$A${PM_HELPER_FIRST_ROW}:$A${PM_HELPER_LAST_ROW},0)+{year_offset}),""))'
    )


def style_results_sheet(ws, result_last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "70AD47"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 32
    style_cells(ws, "A1:U1", fill=NAVY, font_color=WHITE, bold=True, size=15)
    style_cells(ws, "A2:U2", fill=TEAL, font_color=TEXT)
    style_cells(ws, "A4:U4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    style_cells(ws, "V3:V3", fill=TEAL, font_color=TEXT, bold=True, h_align="center")
    style_cells(ws, "V4:V4", fill=NAVY, font_color=WHITE, bold=True, h_align="center")
    set_link(ws, "A3", "Dashboard", "Financial Dashboard")
    set_link(ws, "C3", "Scenario Analysis", ANALYSIS_SHEET)
    set_link(ws, "E3", "Pivot Lab", PIVOT_SHEET)

    ws.column_dimensions["V"].width = 18
    ws["V4"] = "Trend"
    set_header(ws["V4"])
    ws["V3"] = "Quick Path Shape"
    ws["V3"].fill = PatternFill("solid", fgColor=TEAL)
    ws["V3"].font = Font(bold=True)
    ws["V3"].alignment = Alignment(horizontal="center", vertical="center")
    thin_bottom(ws["V3"])

    helper_start = 23
    for idx in range(15):
        col = get_column_letter(helper_start + idx)
        ws.column_dimensions[col].width = 1.1
        ws.column_dimensions[col].hidden = False
        ws.cell(4, helper_start + idx, f"Y{idx + 1}")
        set_header(ws.cell(4, helper_start + idx))

    for row in range(RESULT_FIRST_ROW, result_last_row + 1):
        ws.row_dimensions[row].height = 20
        ws[f"V{row}"].fill = PatternFill("solid", fgColor="F3F7FB")
        thin_bottom(ws[f"V{row}"])
        for idx in range(15):
            col = get_column_letter(helper_start + idx)
            ws[f"{col}{row}"] = liquid_series_formula(row, idx)
            ws[f"{col}{row}"].number_format = MONEY_FMT
            ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=WHITE)
            ws[f"{col}{row}"].font = Font(name="Aptos", color=WHITE, size=8)
            thin_bottom(ws[f"{col}{row}"])

    for merged in list(ws.merged_cells.ranges):
        if str(merged) == "A65:U65":
            ws.unmerge_cells(str(merged))
    clear_row(ws, 65, 1, 21)


def build_pivot_lab_sheet(wb) -> None:
    if PIVOT_SHEET in wb.sheetnames:
        ws = wb[PIVOT_SHEET]
    else:
        ws = wb.create_sheet(PIVOT_SHEET)

    move_sheet_after(wb, PIVOT_SHEET, ANALYSIS_SHEET)
    ws = wb[PIVOT_SHEET]
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "ED7D31"
    ws.freeze_panes = "A7"

    for col, width in {"A": 24, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18}.items():
        ws.column_dimensions[col].width = width

    set_section_band(ws, "A1:G1", "Scenario Pivot Lab", NAVY)
    safe_merge(ws, "A2:G2")
    ws["A2"] = (
        "Compact native Excel pivot views live here. Use this sheet for quick cuts of the normalized "
        "scenario table without crowding the main analysis page."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34

    set_link(ws, "A3", "Scenario Analysis", ANALYSIS_SHEET)
    set_link(ws, "C3", "Scenario Results", SCENARIO_RESULTS)
    set_link(ws, "E3", "Dashboard", "Financial Dashboard")

    set_section_band(ws, "A5:G5", "Pivot Output", PURPLE)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    for row in range(5, 41):
        ws.row_dimensions[row].height = 20
    if not ws["A6"].value:
        ws["A6"] = "Native Excel PivotTable output starts below."
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center")


def reorder_sheets(wb) -> None:
    desired_prefix = [
        "Financial Dashboard",
        ANALYSIS_SHEET,
        PIVOT_SHEET,
        "Model Inputs",
        "Savings Projection",
        "Scenario Lab",
        IC_SHEET,
        PM_SHEET,
        SCENARIO_RESULTS,
        "Tax Assumptions",
    ]
    ordered = [wb[name] for name in desired_prefix if name in wb.sheetnames]
    remaining = [ws for ws in wb.worksheets if ws.title not in desired_prefix]
    wb._sheets = ordered + remaining


def get_sheet_xml_path(workbook_path: Path, sheet_name: str) -> str:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with ZipFile(workbook_path) as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: (
            rel.attrib["Target"].lstrip("/")
            if rel.attrib["Target"].lstrip("/").startswith("xl/")
            else f"xl/{rel.attrib['Target'].lstrip('/')}"
        )
        for rel in rels.findall("rel:Relationship", ns)
    }
    for sheet in workbook.findall("main:sheets/main:sheet", ns):
        if sheet.attrib["name"] == sheet_name:
            rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            return rel_map[rel_id]
    raise KeyError(f"Sheet not found in workbook package: {sheet_name}")


def restore_sheet_sparklines(source_workbook: Path, target_workbook: Path, sheet_name: str) -> None:
    source_sheet_path = get_sheet_xml_path(source_workbook, sheet_name)
    target_sheet_path = get_sheet_xml_path(target_workbook, sheet_name)

    with ZipFile(source_workbook) as src_zip:
        source_sheet_xml = src_zip.read(source_sheet_path).decode("utf-8")

    with ZipFile(target_workbook) as dst_zip:
        payloads = {name: dst_zip.read(name) for name in dst_zip.namelist()}

    target_sheet_xml = payloads[target_sheet_path].decode("utf-8")
    source_open_tag = re.search(r"<worksheet\b[^>]*>", source_sheet_xml, re.S)
    target_open_tag = re.search(r"<worksheet\b[^>]*>", target_sheet_xml, re.S)
    ext_block = re.search(r"<extLst>.*?</extLst>", source_sheet_xml, re.S)
    if not (source_open_tag and target_open_tag and ext_block):
        raise RuntimeError(f"Could not locate sparkline metadata for {sheet_name}.")

    target_sheet_xml = target_sheet_xml.replace(target_open_tag.group(0), source_open_tag.group(0), 1)
    target_sheet_xml = re.sub(r"<extLst>.*?</extLst>", "", target_sheet_xml, flags=re.S)
    target_sheet_xml = target_sheet_xml.replace("</worksheet>", f"{ext_block.group(0)}</worksheet>")
    payloads[target_sheet_path] = target_sheet_xml.encode("utf-8")

    rebuilt = target_workbook.with_suffix(".sparkline-restored.xlsx")
    with ZipFile(rebuilt, "w", compression=ZIP_DEFLATED) as out_zip:
        for name, data in payloads.items():
            out_zip.writestr(name, data)
    rebuilt.replace(target_workbook)


def main() -> None:
    backup_path = Path("tmp/spreadsheets/Net Worth.before-presentation-pass.xlsx")
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(WORKBOOK, backup_path)

    wb = load_workbook(WORKBOOK)
    result_last_row = get_result_last_row(wb)

    build_pivot_lab_sheet(wb)
    reorder_sheets(wb)

    style_dashboard(wb["Financial Dashboard"])
    style_model_inputs_sheet(wb["Model Inputs"])
    style_savings_projection_sheet(wb["Savings Projection"])
    style_scenario_lab_sheet(wb["Scenario Lab"])
    style_ic_switch_sheet(wb[IC_SHEET])
    style_pm_switch_sheet(wb[PM_SHEET])
    style_analysis_sheet(wb[ANALYSIS_SHEET], result_last_row)
    style_results_sheet(wb[SCENARIO_RESULTS], result_last_row)
    style_tax_sheet(wb["Tax Assumptions"])

    replace_defined_name(wb, "AnalysisPrimaryResultRow", f"'{ANALYSIS_SHEET}'!$K$8")
    replace_defined_name(wb, "AnalysisComparisonResultRow", f"'{ANALYSIS_SHEET}'!$K$10")
    replace_defined_name(wb, "DashboardScenarioResultRow", "'Financial Dashboard'!$P$104")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    with TemporaryDirectory() as tmp_dir:
        tmp_workbook = Path(tmp_dir) / WORKBOOK.name
        wb.save(tmp_workbook)
        restore_sheet_sparklines(WORKBOOK, tmp_workbook, SCENARIO_RESULTS)
        shutil.copy2(tmp_workbook, WORKBOOK)


if __name__ == "__main__":
    main()
