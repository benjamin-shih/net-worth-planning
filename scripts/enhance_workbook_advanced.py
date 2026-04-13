from __future__ import annotations

import argparse

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

from enhance_workbook_low_risk import (
    BLUE,
    GRAY,
    GREEN,
    GREEN_FONT,
    INT_FMT,
    MONEY_FMT,
    MULT_FMT,
    NAVY,
    PURPLE,
    RED,
    RED_FONT,
    SCENARIO_RESULTS,
    TARGET_5M_50_DOWN,
    TARGET_5M_ALL_CASH,
    TARGET_7M_50_DOWN,
    TARGET_7M_ALL_CASH,
    TARGET_STATUS,
    TEAL,
    WHITE,
    WORKBOOK,
    replace_defined_name,
    safe_merge,
    set_header,
    set_label,
    thin_bottom,
)


ANALYSIS_SHEET = "Scenario Analysis"
IC_SHEET = "IC Switch Scenarios"
PM_SHEET = "PM After Switch"
IC_HELPER_FIRST_ROW = 36
IC_HELPER_LAST_ROW = 170
PM_HELPER_FIRST_ROW = 51
PM_HELPER_LAST_ROW = 350
SUPPORT_SCENARIO_FIRST_ROW = 4
SUPPORT_SCENARIO_LAST_ROW = 32
SUPPORT_HORIZON_FIRST_ROW = 4
SUPPORT_HORIZON_LAST_ROW = 5
SUPPORT_METRIC_FIRST_ROW = 4
SUPPORT_METRIC_LAST_ROW = 7
CHART_FIRST_ROW = 42
CHART_LAST_ROW = 56


def get_result_last_row(wb) -> int:
    table = wb[SCENARIO_RESULTS].tables["tblScenarioResults"]
    _, _, _, max_row = range_boundaries(table.ref)
    return max_row


def reset_analysis_sheet(wb):
    if ANALYSIS_SHEET in wb.sheetnames:
        del wb[ANALYSIS_SHEET]
    insert_at = wb.sheetnames.index(SCENARIO_RESULTS) + 1
    return wb.create_sheet(ANALYSIS_SHEET, insert_at)


def clear_validations(ws, targets: list[str]) -> None:
    for dv in list(ws.data_validations.dataValidation):
        sqref = str(dv.sqref)
        if any(target in sqref for target in targets):
            ws.data_validations.dataValidation.remove(dv)


def analysis_result_lookup(col: str, pointer_cell: str, result_last_row: int) -> str:
    return f'=IFERROR(INDEX(\'{SCENARIO_RESULTS}\'!${col}$5:${col}${result_last_row},{pointer_cell}),"")'


def numeric_delta_formula(primary: str, comparison: str) -> str:
    return f'=IF(AND(ISNUMBER({primary}),ISNUMBER({comparison})),{primary}-{comparison},"")'


def helper_offset_formula(path_cell: str, scenario_cell: str) -> str:
    return (
        f'=IF({path_cell}="","",'
        f'IF({path_cell}="IC Switch",'
        f'IFERROR(MATCH({scenario_cell},\'{IC_SHEET}\'!$A${IC_HELPER_FIRST_ROW}:$A${IC_HELPER_LAST_ROW},0),""),'
        f'IFERROR(MATCH({scenario_cell},\'{PM_SHEET}\'!$A${PM_HELPER_FIRST_ROW}:$A${PM_HELPER_LAST_ROW},0),"")))'
    )


def metric_series_formula(path_cell: str, offset_cell: str, metric_index_cell: str, year_cell: str) -> str:
    ic_terms = ",".join(
        [
            f"INDEX('{IC_SHEET}'!$L${IC_HELPER_FIRST_ROW}:$L${IC_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{IC_SHEET}'!$V${IC_HELPER_FIRST_ROW}:$V${IC_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{IC_SHEET}'!$W${IC_HELPER_FIRST_ROW}:$W${IC_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{IC_SHEET}'!$X${IC_HELPER_FIRST_ROW}:$X${IC_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
        ]
    )
    pm_terms = ",".join(
        [
            f"INDEX('{PM_SHEET}'!$T${PM_HELPER_FIRST_ROW}:$T${PM_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{PM_SHEET}'!$AD${PM_HELPER_FIRST_ROW}:$AD${PM_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{PM_SHEET}'!$AE${PM_HELPER_FIRST_ROW}:$AE${PM_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
            f"INDEX('{PM_SHEET}'!$AF${PM_HELPER_FIRST_ROW}:$AF${PM_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)",
        ]
    )
    return (
        f'=IF({offset_cell}="","",'
        f'IF({path_cell}="IC Switch",CHOOSE({metric_index_cell},{ic_terms}),'
        f'CHOOSE({metric_index_cell},{pm_terms})))'
    )


def liquid_nw_formula(path_cell: str, offset_cell: str, year_cell: str) -> str:
    return (
        f'=IF({offset_cell}="","",'
        f'IF({path_cell}="IC Switch",'
        f"INDEX('{IC_SHEET}'!$X${IC_HELPER_FIRST_ROW}:$X${IC_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1),"
        f"INDEX('{PM_SHEET}'!$AF${PM_HELPER_FIRST_ROW}:$AF${PM_HELPER_LAST_ROW},{offset_cell}+{year_cell}-1)))"
    )


def style_section_header(ws, cell_range: str, value: str) -> None:
    safe_merge(ws, cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=PURPLE)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def populate_support_lists(ws) -> None:
    style_section_header(ws, "X1:AB1", "Support Lists")
    headers = ["Scenario", "Path", "Summary Row", "Helper Sheet", "Helper Start Row"]
    for idx, header in enumerate(headers, 24):
        set_header(ws.cell(2, idx))
        ws.cell(2, idx, header)

    out_row = SUPPORT_SCENARIO_FIRST_ROW
    for source_row in range(22, 31):
        ws[f"X{out_row}"] = f"='{IC_SHEET}'!$A${source_row}"
        ws[f"Y{out_row}"] = "IC Switch"
        ws[f"Z{out_row}"] = source_row
        ws[f"AA{out_row}"] = IC_SHEET
        ws[f"AB{out_row}"] = f'=MATCH($X{out_row},\'{IC_SHEET}\'!$A${IC_HELPER_FIRST_ROW}:$A${IC_HELPER_LAST_ROW},0)'
        out_row += 1

    for source_row in range(22, 42):
        ws[f"X{out_row}"] = f"='{PM_SHEET}'!$A${source_row}"
        ws[f"Y{out_row}"] = "PM After Switch"
        ws[f"Z{out_row}"] = source_row
        ws[f"AA{out_row}"] = PM_SHEET
        ws[f"AB{out_row}"] = f'=MATCH($X{out_row},\'{PM_SHEET}\'!$A${PM_HELPER_FIRST_ROW}:$A${PM_HELPER_LAST_ROW},0)'
        out_row += 1

    style_section_header(ws, "AD1:AE1", "Selector Options")
    ws["AD2"] = "Horizon"
    ws["AE2"] = "Metric"
    set_header(ws["AD2"])
    set_header(ws["AE2"])
    ws["AD4"] = "Y10"
    ws["AD5"] = "Y15"
    ws["AE4"] = "Scenario Cash Gross"
    ws["AE5"] = "Taxable Balance"
    ws["AE6"] = "Retirement Balance"
    ws["AE7"] = "Liquid Net Worth"

    for cell_range in [f"X{SUPPORT_SCENARIO_FIRST_ROW}:X{SUPPORT_SCENARIO_LAST_ROW}", "AD4:AD5", "AE4:AE7"]:
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=TEAL)
                thin_bottom(cell)


def add_analysis_controls(ws, result_last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "5B9BD5"

    widths = {
        "A": 22,
        "B": 28,
        "C": 14,
        "D": 14,
        "E": 4,
        "F": 18,
        "G": 20,
        "H": 4,
        "I": 4,
        "J": 20,
        "K": 32,
        "L": 4,
        "M": 18,
        "N": 14,
        "O": 4,
        "P": 4,
        "X": 28,
        "Y": 18,
        "Z": 12,
        "AA": 18,
        "AB": 14,
        "AD": 14,
        "AE": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    safe_merge(ws, "A1:P1")
    ws["A1"] = "Scenario Analysis"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    safe_merge(ws, "A2:P2")
    ws["A2"] = (
        "Compare specific switch outcomes quickly, benchmark them against staying at Jump, view "
        "liquid-net-worth matrices at the chosen horizon, and inspect compact Y1-Y15 trajectory charts."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A3"] = "Back to Financial Dashboard"
    ws["A3"].hyperlink = "#'Financial Dashboard'!A1"
    ws["A3"].style = "Hyperlink"

    style_section_header(ws, "A4:G4", "Controls")
    for label_cell, value_cell, value in [
        ("A5", "B5", "3Y 2x"),
        ("A6", "B6", "3Y -> PM Y8 Base"),
        ("F5", "G5", "Y15"),
        ("F6", "G6", "Liquid Net Worth"),
    ]:
        ws[label_cell] = {
            "A5": "Primary scenario",
            "A6": "Comparison scenario",
            "F5": "Summary horizon",
            "F6": "Chart metric",
        }[label_cell]
        set_label(ws[label_cell])
        ws[value_cell] = value
        ws[value_cell].fill = PatternFill("solid", fgColor=BLUE)
        ws[value_cell].font = Font(color="0000FF")
        ws[value_cell].alignment = Alignment(horizontal="left", vertical="center")
        thin_bottom(ws[value_cell])

    style_section_header(ws, "J4:N4", "Selector Helpers")
    helper_labels = {
        "J5": "Primary path",
        "J6": "Comparison path",
        "J7": "Primary result key",
        "J8": "Primary result row",
        "J9": "Comparison result key",
        "J10": "Comparison result row",
        "J11": "Baseline result key",
        "J12": "Baseline result row",
        "M5": "Primary helper offset",
        "M6": "Comparison helper offset",
        "M7": "Metric index",
    }
    for cell_ref, label in helper_labels.items():
        ws[cell_ref] = label
        set_label(ws[cell_ref])

    ws["K5"] = '=IFERROR(INDEX($Y$4:$Y$32,MATCH($B$5,$X$4:$X$32,0)),"")'
    ws["K6"] = '=IFERROR(INDEX($Y$4:$Y$32,MATCH($B$6,$X$4:$X$32,0)),"")'
    ws["K7"] = '=$K$5&" | "&$B$5&" | "&$G$5'
    ws["K8"] = f'=IFERROR(MATCH($K$7,\'{SCENARIO_RESULTS}\'!$A$5:$A${result_last_row},0),"")'
    ws["K9"] = '=$K$6&" | "&$B$6&" | "&$G$5'
    ws["K10"] = f'=IFERROR(MATCH($K$9,\'{SCENARIO_RESULTS}\'!$A$5:$A${result_last_row},0),"")'
    ws["K11"] = '="Stay at Jump | Base | "&$G$5'
    ws["K12"] = f'=IFERROR(MATCH($K$11,\'{SCENARIO_RESULTS}\'!$A$5:$A${result_last_row},0),"")'
    ws["N5"] = helper_offset_formula("$K$5", "$B$5")
    ws["N6"] = helper_offset_formula("$K$6", "$B$6")
    ws["N7"] = '=IFERROR(MATCH($G$6,$AE$4:$AE$7,0),"")'

    for cell_ref in ("K5", "K6", "K7", "K8", "K9", "K10", "K11", "K12", "N5", "N6", "N7"):
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_bottom(ws[cell_ref])

    for cell_ref in ("K8", "K10", "K12", "N5", "N6", "N7"):
        ws[cell_ref].number_format = INT_FMT

    clear_validations(ws, ["B5", "B6", "G5", "G6"])
    scenario_dv = DataValidation(type="list", formula1="=$X$4:$X$32", allow_blank=False)
    horizon_dv = DataValidation(type="list", formula1="=$AD$4:$AD$5", allow_blank=False)
    metric_dv = DataValidation(type="list", formula1="=$AE$4:$AE$7", allow_blank=False)
    ws.add_data_validation(scenario_dv)
    ws.add_data_validation(horizon_dv)
    ws.add_data_validation(metric_dv)
    scenario_dv.add("B5")
    scenario_dv.add("B6")
    horizon_dv.add("G5")
    metric_dv.add("G6")


def add_summary_table(ws, result_last_row: int) -> None:
    style_section_header(ws, "A12:G12", "Selected Horizon Comparison")
    headers = {
        "A13": "Metric",
        "B13": "Primary",
        "C13": "Comparison",
        "D13": "Stay at Jump (Base)",
        "E13": "Primary - Comparison",
        "F13": "Primary - Jump Base",
        "G13": "Comparison - Jump Base",
    }
    for cell_ref, label in headers.items():
        ws[cell_ref] = label
        set_header(ws[cell_ref])

    labels = [
        ("A14", "Path"),
        ("A15", "Scenario"),
        ("A16", "Horizon"),
        ("A17", "Cash / Gross Comp"),
        ("A18", "Taxable Liquid"),
        ("A19", "Retirement"),
        ("A20", "Liquid Net Worth"),
        ("A21", TARGET_5M_ALL_CASH),
        ("A22", TARGET_5M_50_DOWN),
        ("A23", TARGET_7M_ALL_CASH),
        ("A24", TARGET_7M_50_DOWN),
        ("A25", "First $5M Cash"),
        ("A26", "First $5M 50%"),
        ("A27", "First $7M Cash"),
        ("A28", "First $7M 50%"),
        ("A29", TARGET_STATUS),
    ]
    for cell_ref, label in labels:
        ws[cell_ref] = label
        set_label(ws[cell_ref])

    lookup_cols = {
        14: "B",
        15: "C",
        16: "H",
        17: "I",
        18: "J",
        19: "K",
        20: "L",
        21: "M",
        22: "N",
        23: "O",
        24: "P",
        25: "R",
        26: "S",
        27: "T",
        28: "U",
        29: "Q",
    }
    for row, col in lookup_cols.items():
        ws[f"B{row}"] = analysis_result_lookup(col, "$K$8", result_last_row)
        ws[f"C{row}"] = analysis_result_lookup(col, "$K$10", result_last_row)
        ws[f"D{row}"] = analysis_result_lookup(col, "$K$12", result_last_row)
        thin_bottom(ws[f"B{row}"])
        thin_bottom(ws[f"C{row}"])
        thin_bottom(ws[f"D{row}"])
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in range(17, 29):
        ws[f"E{row}"] = numeric_delta_formula(f"B{row}", f"C{row}")
        ws[f"F{row}"] = numeric_delta_formula(f"B{row}", f"D{row}")
        ws[f"G{row}"] = numeric_delta_formula(f"C{row}", f"D{row}")
        for col in ("E", "F", "G"):
            thin_bottom(ws[f"{col}{row}"])
            ws[f"{col}{row}"].alignment = Alignment(horizontal="right", vertical="center")

    for row in range(17, 25):
        for col in ("B", "C", "D", "E", "F", "G"):
            ws[f"{col}{row}"].number_format = MONEY_FMT
    for row in range(25, 29):
        for col in ("B", "C", "D", "E", "F", "G"):
            ws[f"{col}{row}"].number_format = INT_FMT

    green_fill = PatternFill("solid", fgColor=GREEN)
    red_fill = PatternFill("solid", fgColor=RED)
    green_font = Font(color=GREEN_FONT)
    red_font = Font(color=RED_FONT)
    for cell_range in ("E17:G24",):
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill, font=green_font),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font),
        )


def add_liquid_nw_matrices(ws, result_last_row: int) -> None:
    del result_last_row
    style_section_header(ws, "A32:D32", '="IC Switch Liquid Net Worth ("&$G$5&")"')
    for cell_ref, label in {"B33": "2x", "C33": "3x", "D33": "4x"}.items():
        ws[cell_ref] = label
        set_header(ws[cell_ref])
    for row, yoe in zip(range(34, 37), ("3Y", "4Y", "5Y")):
        ws[f"A{row}"] = yoe
        set_label(ws[f"A{row}"])
    for row in range(34, 37):
        for col in ("B", "C", "D"):
            ws[f"{col}{row}"] = (
                '=SUMIFS('
                'tblScenarioResults[Liquid Net Worth],'
                'tblScenarioResults[Path],"IC Switch",'
                f'tblScenarioResults[Scenario],$A{row}&" "&{col}$33,'
                'tblScenarioResults[Horizon],$G$5)'
            )
            ws[f"{col}{row}"].number_format = MONEY_FMT
            thin_bottom(ws[f"{col}{row}"])

    style_section_header(ws, "J32:N32", '="PM After Switch Liquid Net Worth ("&$G$5&")"')
    for cell_ref, label in {"K33": "Starter", "L33": "Base", "M33": "Upside", "N33": "Tail"}.items():
        ws[cell_ref] = label
        set_header(ws[cell_ref])
    pm_row_labels = {
        34: "3Y -> PM Y7",
        35: "3Y -> PM Y8",
        36: "4Y -> PM Y7",
        37: "4Y -> PM Y8",
        38: "5Y -> PM Y8",
    }
    for row, label in pm_row_labels.items():
        ws[f"J{row}"] = label
        set_label(ws[f"J{row}"])
    for row in range(34, 39):
        for col in ("K", "L", "M", "N"):
            ws[f"{col}{row}"] = (
                '=SUMIFS('
                'tblScenarioResults[Liquid Net Worth],'
                'tblScenarioResults[Path],"PM After Switch",'
                f'tblScenarioResults[Scenario],$J{row}&" "&{col}$33,'
                'tblScenarioResults[Horizon],$G$5)'
            )
            ws[f"{col}{row}"].number_format = MONEY_FMT
            thin_bottom(ws[f"{col}{row}"])

    ws.conditional_formatting.add(
        "B34:D36",
        ColorScaleRule(
            start_type="min",
            start_color="F4CCCC",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="D9EAD3",
        ),
    )
    ws.conditional_formatting.add(
        "K34:N38",
        ColorScaleRule(
            start_type="min",
            start_color="F4CCCC",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="D9EAD3",
        ),
    )


def add_chart_data_block(ws) -> None:
    style_section_header(ws, "X38:AE38", "Chart Data")
    headers = {
        "X41": "Projection Year",
        "Y41": "Calendar Year",
        "Z41": "Primary Selected Metric",
        "AA41": "Comparison Selected Metric",
        "AB41": "Stay at Jump Selected Metric",
        "AC41": "Primary Liquid Net Worth",
        "AD41": "Comparison Liquid Net Worth",
        "AE41": "Stay at Jump Liquid Net Worth",
    }
    for cell_ref, label in headers.items():
        ws[cell_ref] = label
        set_header(ws[cell_ref])

    for row in range(CHART_FIRST_ROW, CHART_LAST_ROW + 1):
        ws[f"X{row}"] = row - CHART_FIRST_ROW + 1
        ws[f"Y{row}"] = f"='Model Inputs'!$B$22-1+X{row}"
        ws[f"Z{row}"] = metric_series_formula("$K$5", "$N$5", "$N$7", f"X{row}")
        ws[f"AA{row}"] = metric_series_formula("$K$6", "$N$6", "$N$7", f"X{row}")
        ws[f"AB{row}"] = (
            f'=CHOOSE($N$7,'
            f"INDEX('Savings Projection'!$B$19:$B$33,X{row}),"
            f"INDEX('Savings Projection'!$AE$19:$AE$33,X{row}),"
            f"INDEX('Savings Projection'!$AF$19:$AF$33,X{row})+INDEX('Savings Projection'!$AG$19:$AG$33,X{row})+INDEX('Savings Projection'!$AH$19:$AH$33,X{row}),"
            f"INDEX('Savings Projection'!$AI$19:$AI$33,X{row}))"
        )
        ws[f"AC{row}"] = liquid_nw_formula("$K$5", "$N$5", f"X{row}")
        ws[f"AD{row}"] = liquid_nw_formula("$K$6", "$N$6", f"X{row}")
        ws[f"AE{row}"] = f"=INDEX('Savings Projection'!$AI$19:$AI$33,X{row})"

        ws[f"X{row}"].number_format = INT_FMT
        ws[f"Y{row}"].number_format = INT_FMT
        for col in ("Z", "AA", "AB", "AC", "AD", "AE"):
            ws[f"{col}{row}"].number_format = MONEY_FMT
        for col in ("X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"):
            thin_bottom(ws[f"{col}{row}"])


def build_analysis_sheet(wb, result_last_row: int):
    ws = reset_analysis_sheet(wb)
    populate_support_lists(ws)
    add_analysis_controls(ws, result_last_row)
    add_summary_table(ws, result_last_row)
    add_liquid_nw_matrices(ws, result_last_row)
    add_chart_data_block(ws)
    ws.row_dimensions[2].height = 30
    for row in range(4, 11):
        ws.row_dimensions[row].height = 22
    return ws


def configure_line_chart(chart: LineChart, title: str) -> None:
    chart.title = title
    chart.style = 10
    chart.height = 7.2
    chart.width = 9.6
    chart.legend.position = "b"
    chart.y_axis.numFmt = '$0.0,,"M"'
    chart.x_axis.title = "Calendar Year"


def add_analysis_charts(ws) -> None:
    style_section_header(ws, "A40:G40", "Trajectory Charts")
    chart_metric = LineChart()
    configure_line_chart(chart_metric, "Selected Metric Trend")
    metric_data = Reference(ws, min_col=26, min_row=41, max_col=28, max_row=56)
    metric_cats = Reference(ws, min_col=25, min_row=42, max_row=56)
    chart_metric.add_data(metric_data, titles_from_data=True)
    chart_metric.set_categories(metric_cats)
    ws.add_chart(chart_metric, "A41")

    chart_nw = LineChart()
    configure_line_chart(chart_nw, "Liquid Net Worth Trend")
    nw_data = Reference(ws, min_col=29, min_row=41, max_col=31, max_row=56)
    nw_cats = Reference(ws, min_col=25, min_row=42, max_row=56)
    chart_nw.add_data(nw_data, titles_from_data=True)
    chart_nw.set_categories(nw_cats)
    ws.add_chart(chart_nw, "I41")


def apply_named_cleanup(wb, result_last_row: int) -> None:
    ws = wb["Financial Dashboard"]
    ws["O104"] = "Selector row"
    set_label(ws["O104"])
    ws["P104"] = f'=MATCH($B$103,\'{SCENARIO_RESULTS}\'!$A$5:$A${result_last_row},0)'
    ws["P104"].number_format = INT_FMT
    thin_bottom(ws["P104"])
    replace_defined_name(wb, "DashboardScenarioResultKey", "'Financial Dashboard'!$B$103")
    replace_defined_name(wb, "DashboardScenarioResultRow", "'Financial Dashboard'!$P$104")
    replace_defined_name(wb, "DashboardSelectedProjectionRow", "'Financial Dashboard'!$J$65")
    replace_defined_name(wb, "AnalysisPrimaryScenario", f"'{ANALYSIS_SHEET}'!$B$5")
    replace_defined_name(wb, "AnalysisComparisonScenario", f"'{ANALYSIS_SHEET}'!$B$6")
    replace_defined_name(wb, "AnalysisSelectedHorizon", f"'{ANALYSIS_SHEET}'!$G$5")
    replace_defined_name(wb, "AnalysisSelectedMetric", f"'{ANALYSIS_SHEET}'!$G$6")
    replace_defined_name(wb, "AnalysisPrimaryResultRow", f"'{ANALYSIS_SHEET}'!$K$8")
    replace_defined_name(wb, "AnalysisComparisonResultRow", f"'{ANALYSIS_SHEET}'!$K$10")
    replace_defined_name(wb, "AnalysisMetricIndex", f"'{ANALYSIS_SHEET}'!$N$7")

    ws["J65"] = "=MAX(1,MIN(ROUND('Model Inputs'!$B$15,0),ROWS('Savings Projection'!$A$19:$A$58)))"
    ws["O65"] = "=INDEX('Savings Projection'!$C$19:$C$58,DashboardSelectedProjectionRow)"
    ws["J66"] = "=INDEX('Savings Projection'!$B$19:$B$58,DashboardSelectedProjectionRow)"
    ws["J67"] = "=INDEX('Savings Projection'!$AJ$19:$AJ$58,DashboardSelectedProjectionRow)"
    ws["J68"] = "=INDEX('Savings Projection'!$AK$19:$AK$58,DashboardSelectedProjectionRow)"
    ws["J69"] = "=INDEX('Savings Projection'!$X$19:$X$58,DashboardSelectedProjectionRow)"

    lookup = lambda col: f"=IFERROR(INDEX('{SCENARIO_RESULTS}'!${col}$5:${col}${result_last_row},DashboardScenarioResultRow),\"\")"
    ws["B105"] = lookup("B")
    ws["B106"] = lookup("C")
    ws["B107"] = lookup("H")
    ws["B108"] = lookup("D")
    ws["B109"] = lookup("E")
    ws["B110"] = lookup("F")
    ws["B111"] = lookup("G")
    ws["G105"] = lookup("I")
    ws["J105"] = lookup("J")
    ws["M105"] = lookup("K")
    ws["P105"] = lookup("L")
    ws["G107"] = lookup("M")
    ws["J107"] = lookup("N")
    ws["M107"] = lookup("O")
    ws["P107"] = lookup("P")
    ws["G109"] = lookup("R")
    ws["J109"] = lookup("S")
    ws["M109"] = lookup("T")
    ws["P109"] = lookup("U")
    ws["F111"] = lookup("Q")

    safe_merge(ws, "A114:P114")
    for cell_ref, label, sheet in [
        ("A115", "Dashboard", "Financial Dashboard"),
        ("C115", "Model Inputs", "Model Inputs"),
        ("E115", "Savings Projection", "Savings Projection"),
        ("G115", "Scenario Lab", "Scenario Lab"),
        ("I115", "IC Switch", IC_SHEET),
        ("K115", "PM Switch", PM_SHEET),
        ("M115", "Analysis", ANALYSIS_SHEET),
        ("O115", "Scenario Results", SCENARIO_RESULTS),
        ("A116", "Tax", "Tax Assumptions"),
    ]:
        ws[cell_ref] = label
        ws[cell_ref].hyperlink = f"#'{sheet}'!A1"
        ws[cell_ref].style = "Hyperlink"
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply the advanced workbook presentation pass.")
    parser.add_argument(
        "--phase",
        type=int,
        choices=(1, 2, 3),
        default=3,
        help="1=create the analysis sheet; 2=also add charts; 3=also add named-formula cleanup.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    wb = load_workbook(WORKBOOK)
    result_last_row = get_result_last_row(wb)
    ws = build_analysis_sheet(wb, result_last_row)
    if args.phase >= 2:
        add_analysis_charts(ws)
    if args.phase >= 3:
        apply_named_cleanup(wb, result_last_row)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
