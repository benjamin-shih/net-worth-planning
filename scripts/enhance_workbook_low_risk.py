from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


WORKBOOK = Path("Net Worth.xlsx")
SCENARIO_RESULTS = "Scenario Results"
RESULT_FIRST_ROW = 5

MONEY_FMT = '$#,##0;[Red]($#,##0);-'
MULT_FMT = '0.0x'
INT_FMT = '0'

NAVY = "1F4E78"
BLUE = "D9EAF7"
TEAL = "DDEBF7"
GREEN = "E2F0D9"
GREEN_FONT = "006100"
RED = "F4CCCC"
RED_FONT = "9C0006"
ORANGE = "FCE4D6"
PURPLE = "E4DFEC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"


def sheet_ref(sheet: str, cell: str) -> str:
    col, row = coordinate_from_string(cell)
    return f"='{sheet}'!${col}${row}"


def replace_table(ws, name: str, ref: str, style: str) -> None:
    if name in ws.tables:
        del ws.tables[name]
    if ws.auto_filter.ref == ref:
        ws.auto_filter.ref = None

    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_label(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=GRAY)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def thin_bottom(cell) -> None:
    side = Side(style="thin", color="A6A6A6")
    cell.border = Border(bottom=side)


def safe_merge(ws, cell_range: str) -> None:
    for merged in list(ws.merged_cells.ranges):
        if str(merged) == cell_range:
            return
    ws.merge_cells(cell_range)


def unmerge_intersecting_rows(ws, min_row: int, max_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        _, merged_min_row, _, merged_max_row = range_boundaries(str(merged))
        if merged_min_row <= max_row and merged_max_row >= min_row:
            ws.unmerge_cells(str(merged))


def style_result_sheet(ws, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "70AD47"

    widths = {
        "A": 34,
        "B": 18,
        "C": 28,
        "D": 12,
        "E": 10,
        "F": 15,
        "G": 14,
        "H": 10,
        "I": 16,
        "J": 16,
        "K": 14,
        "L": 16,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 15,
        "Q": 30,
        "R": 15,
        "S": 15,
        "T": 15,
        "U": 15,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    safe_merge(ws, "A1:U1")
    ws["A1"] = "Scenario Results"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    safe_merge(ws, "A2:U2")
    ws["A2"] = (
        "Normalized view of IC switch and PM-after-switch outcomes. "
        "Use filters or the dashboard selector instead of scanning each scenario sheet separately."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    for cell in ws[4]:
        set_header(cell)

    for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=1, max_col=21):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_bottom(cell)

    for col in ("D", "F"):
        for cell in ws[f"{col}5:{col}{last_row}"]:
            cell[0].number_format = INT_FMT
            cell[0].alignment = Alignment(horizontal="center")
    for cell in ws[f"E5:E{last_row}"]:
        cell[0].number_format = MULT_FMT
        cell[0].alignment = Alignment(horizontal="center")
    for col in ("I", "J", "K", "L", "M", "N", "O", "P"):
        for cell in ws[f"{col}5:{col}{last_row}"]:
            cell[0].number_format = MONEY_FMT
    for col in ("H", "R", "S", "T", "U"):
        for cell in ws[f"{col}5:{col}{last_row}"]:
            cell[0].alignment = Alignment(horizontal="center")


def add_summary_tables(wb) -> None:
    replace_table(wb["IC Switch Scenarios"], "tblICSwitchSummary", "A21:Z30", "TableStyleMedium2")
    replace_table(wb["PM After Switch"], "tblPMAfterSwitchSummary", "A21:AF41", "TableStyleMedium4")


def add_scenario_results(wb) -> int:
    if SCENARIO_RESULTS in wb.sheetnames:
        del wb[SCENARIO_RESULTS]

    ws = wb.create_sheet(SCENARIO_RESULTS, wb.sheetnames.index("PM After Switch") + 1)

    headers = [
        "Result Key",
        "Path",
        "Scenario",
        "Switch After YOE",
        "IC Bump",
        "Requested PM Start YOE",
        "PM Case",
        "Horizon",
        "Cash/Gross Comp",
        "Taxable Liquid",
        "Retirement",
        "Liquid Net Worth",
        "vs $5M Cash",
        "vs $5M 50%",
        "vs $7M Cash",
        "vs $7M 50%",
        "Read",
        "First $5M Cash",
        "First $5M 50%",
        "First $7M Cash",
        "First $7M 50%",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)

    rows: list[list[str | int | float | None]] = []
    ic = "IC Switch Scenarios"
    for source_row in range(22, 31):
        for horizon, source_cols in (
            ("Y10", ("E", "F", "G", "H", "I", "J", "K", "L", "M")),
            ("Y15", ("N", "O", "P", "Q", "R", "S", "T", "U", "V")),
        ):
            rows.append(
                [
                    None,
                    "IC Switch",
                    sheet_ref(ic, f"A{source_row}"),
                    sheet_ref(ic, f"B{source_row}"),
                    sheet_ref(ic, f"C{source_row}"),
                    "",
                    "",
                    horizon,
                    *[sheet_ref(ic, f"{col}{source_row}") for col in source_cols],
                    sheet_ref(ic, f"W{source_row}"),
                    sheet_ref(ic, f"X{source_row}"),
                    sheet_ref(ic, f"Y{source_row}"),
                    sheet_ref(ic, f"Z{source_row}"),
                ]
            )

    pm = "PM After Switch"
    for source_row in range(22, 42):
        for horizon, source_cols in (
            ("Y10", ("K", "L", "M", "N", "O", "P", "Q", "R", "S")),
            ("Y15", ("T", "U", "V", "W", "X", "Y", "Z", "AA", "AB")),
        ):
            rows.append(
                [
                    None,
                    "PM After Switch",
                    sheet_ref(pm, f"A{source_row}"),
                    sheet_ref(pm, f"B{source_row}"),
                    sheet_ref(pm, f"C{source_row}"),
                    sheet_ref(pm, f"D{source_row}"),
                    sheet_ref(pm, f"E{source_row}"),
                    horizon,
                    *[sheet_ref(pm, f"{col}{source_row}") for col in source_cols],
                    sheet_ref(pm, f"AC{source_row}"),
                    sheet_ref(pm, f"AD{source_row}"),
                    sheet_ref(pm, f"AE{source_row}"),
                    sheet_ref(pm, f"AF{source_row}"),
                ]
            )

    for row_idx, row_values in enumerate(rows, RESULT_FIRST_ROW):
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 1, f'=$B{row_idx}&" | "&$C{row_idx}&" | "&$H{row_idx}')

    last_row = RESULT_FIRST_ROW + len(rows) - 1
    style_result_sheet(ws, last_row)
    replace_table(ws, "tblScenarioResults", f"A4:U{last_row}", "TableStyleMedium13")

    safe_merge(ws, "A65:U65")
    ws["A65"] = "Next-Pass Enhancements"
    ws["A65"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["A65"].font = Font(bold=True)
    ws["A65"].alignment = Alignment(horizontal="left")

    notes = [
        "PivotTables and slicers over tblScenarioResults, preferably created with native Excel automation after the normalized table settles.",
        "Sparklines or compact trajectory charts for scenario Y1-Y15 paths, using visible helper rows as the source.",
        "Named formulas or LET/LAMBDA cleanup for repeated phase-gating logic after a separate formula audit.",
        "Power Query only if external data imports become part of the planning workflow.",
    ]
    for row_idx, note in enumerate(notes, 66):
        safe_merge(ws, f"A{row_idx}:U{row_idx}")
        ws.cell(row_idx, 1, note)
        ws.cell(row_idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = 28

    return last_row


def replace_defined_name(wb, name: str, target: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=target))


def add_dashboard_selector(wb, result_last_row: int) -> None:
    ws = wb["Financial Dashboard"]
    ws.sheet_view.showGridLines = False

    unmerge_intersecting_rows(ws, 100, 122)
    for row in range(100, 123):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            cell.value = None
            cell._style = copy(ws["A99"]._style)
            cell.hyperlink = None
            cell.comment = None

    safe_merge(ws, "A100:P100")
    ws["A100"] = "Career Scenario Selector"
    ws["A100"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A100"].font = Font(color=WHITE, bold=True, size=12)
    ws["A100"].alignment = Alignment(horizontal="left", vertical="center")

    safe_merge(ws, "A101:P101")
    ws["A101"] = (
        "Pick a scenario result once; the dashboard pulls the linked Y10 or Y15 readout "
        "from the normalized Scenario Results table."
    )
    ws["A101"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A101"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A103"] = "Selected result"
    set_label(ws["A103"])
    ws["B103"] = "IC Switch | 3Y 2x | Y10"
    ws["B103"].fill = PatternFill("solid", fgColor=BLUE)
    ws["B103"].font = Font(color="0000FF")
    ws["B103"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    safe_merge(ws, "B103:P103")

    lookup_row = f"MATCH($B$103,'{SCENARIO_RESULTS}'!$A$5:$A${result_last_row},0)"
    lookup = lambda col: f"=IFERROR(INDEX('{SCENARIO_RESULTS}'!${col}$5:${col}${result_last_row},{lookup_row}),\"\")"

    left_pairs = [
        ("A105", "Path", "B105", lookup("B")),
        ("A106", "Scenario", "B106", lookup("C")),
        ("A107", "Horizon", "B107", lookup("H")),
        ("A108", "Switch after YOE", "B108", lookup("D")),
        ("A109", "IC bump", "B109", lookup("E")),
        ("A110", "Requested PM start", "B110", lookup("F")),
        ("A111", "PM case", "B111", lookup("G")),
    ]
    for label_cell, label, value_cell, formula in left_pairs:
        ws[label_cell] = label
        set_label(ws[label_cell])
        ws[value_cell] = formula
        ws[value_cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_bottom(ws[value_cell])

    output_pairs = [
        ("F105", "Cash/Gross comp", "G105", lookup("I")),
        ("I105", "Taxable liquid", "J105", lookup("J")),
        ("L105", "Retirement", "M105", lookup("K")),
        ("O105", "Liquid NW", "P105", lookup("L")),
        ("F107", "vs $5M cash", "G107", lookup("M")),
        ("I107", "vs $5M 50%", "J107", lookup("N")),
        ("L107", "vs $7M cash", "M107", lookup("O")),
        ("O107", "vs $7M 50%", "P107", lookup("P")),
        ("F109", "First $5M cash", "G109", lookup("R")),
        ("I109", "First $5M 50%", "J109", lookup("S")),
        ("L109", "First $7M cash", "M109", lookup("T")),
        ("O109", "First $7M 50%", "P109", lookup("U")),
    ]
    for label_cell, label, value_cell, formula in output_pairs:
        ws[label_cell] = label
        set_label(ws[label_cell])
        ws[value_cell] = formula
        ws[value_cell].number_format = MONEY_FMT if value_cell in {"G105", "J105", "M105", "P105", "G107", "J107", "M107", "P107"} else "General"
        ws[value_cell].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin_bottom(ws[value_cell])

    safe_merge(ws, "F111:P111")
    ws["F111"] = lookup("Q")
    ws["F111"].fill = PatternFill("solid", fgColor=TEAL)
    ws["F111"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    safe_merge(ws, "A114:P114")
    ws["A114"] = "Navigation"
    ws["A114"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A114"].font = Font(color=WHITE, bold=True)

    nav = [
        ("A115", "Dashboard", "Financial Dashboard"),
        ("C115", "Model Inputs", "Model Inputs"),
        ("E115", "Savings Projection", "Savings Projection"),
        ("G115", "Scenario Lab", "Scenario Lab"),
        ("I115", "IC Switch", "IC Switch Scenarios"),
        ("K115", "PM Switch", "PM After Switch"),
        ("M115", "Scenario Results", SCENARIO_RESULTS),
        ("O115", "Tax", "Tax Assumptions"),
    ]
    for cell_ref, label, sheet in nav:
        cell = ws[cell_ref]
        cell.value = label
        cell.hyperlink = f"#'{sheet}'!A1"
        cell.style = "Hyperlink"
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(100, 116):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[101].height = 30
    ws.row_dimensions[103].height = 32
    ws.row_dimensions[111].height = 34

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = max(ws.column_dimensions[get_column_letter(col)].width or 0, 13)

    for dv in list(ws.data_validations.dataValidation):
        if "B103" in str(dv.sqref):
            ws.data_validations.dataValidation.remove(dv)
    selector = DataValidation(type="list", formula1="=ScenarioResultKeys", allow_blank=False)
    selector.error = "Select a scenario from the Scenario Results list."
    selector.errorTitle = "Invalid scenario"
    ws.add_data_validation(selector)
    selector.add("B103")

    for cell_ref in ("B108", "B110"):
        ws[cell_ref].number_format = INT_FMT
    ws["B109"].number_format = MULT_FMT


def add_conditional_formatting(wb, result_last_row: int) -> None:
    green_fill = PatternFill("solid", fgColor=GREEN)
    red_fill = PatternFill("solid", fgColor=RED)
    green_font = Font(color=GREEN_FONT)
    red_font = Font(color=RED_FONT)

    def add_delta_rules(ws, ranges: list[str]) -> None:
        for cell_range in ranges:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill, font=green_font),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font),
            )

    def add_text_rules(ws, cell_range: str, top_left: str) -> None:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'ISNUMBER(SEARCH("Short",{top_left}))'], fill=red_fill, font=red_font),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'ISNUMBER(SEARCH("Clears",{top_left}))'], fill=green_fill, font=green_font),
        )

    def add_not_by_rules(ws, cell_range: str, top_left: str) -> None:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'{top_left}="Not by Y15"'], fill=red_fill, font=red_font),
        )

    ic = wb["IC Switch Scenarios"]
    add_delta_rules(ic, ["I22:L30", "R22:U30"])
    add_text_rules(ic, "M22:M30", "M22")
    add_text_rules(ic, "V22:V30", "V22")
    add_not_by_rules(ic, "W22:Z30", "W22")

    pm = wb["PM After Switch"]
    add_delta_rules(pm, ["O22:R41", "X22:AA41"])
    add_text_rules(pm, "S22:S41", "S22")
    add_text_rules(pm, "AB22:AB41", "AB22")
    add_not_by_rules(pm, "AC22:AF41", "AC22")

    sr = wb[SCENARIO_RESULTS]
    add_delta_rules(sr, [f"M5:P{result_last_row}"])
    add_text_rules(sr, f"Q5:Q{result_last_row}", "Q5")
    add_not_by_rules(sr, f"R5:U{result_last_row}", "R5")
    sr.conditional_formatting.add(
        f"L5:L{result_last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F4CCCC",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="D9EAD3",
        ),
    )


def add_sheet_navigation_backlinks(wb) -> None:
    for sheet_name in [
        "Model Inputs",
        "Savings Projection",
        "Scenario Lab",
        "IC Switch Scenarios",
        "PM After Switch",
        SCENARIO_RESULTS,
        "Tax Assumptions",
    ]:
        ws = wb[sheet_name]
        target = ws["A3"]
        if target.value in (None, ""):
            target.value = "Back to Financial Dashboard"
            target.hyperlink = "#'Financial Dashboard'!A1"
            target.style = "Hyperlink"


def main() -> None:
    wb = load_workbook(WORKBOOK)
    add_summary_tables(wb)
    result_last_row = add_scenario_results(wb)
    replace_defined_name(wb, "ScenarioResultKeys", f"'{SCENARIO_RESULTS}'!$A$5:$A${result_last_row}")
    add_dashboard_selector(wb, result_last_row)
    add_conditional_formatting(wb, result_last_row)
    add_sheet_navigation_backlinks(wb)

    # Recalculate in Excel after saving through openpyxl.
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
